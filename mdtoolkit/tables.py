"""
Markdown table extraction and export to CSV / XLSX.
"""

import re
import csv
from pathlib import Path
from mdtoolkit.colors import ok, warn, err


def extract_tables(md_text: str) -> list:
    """
    Parse all pipe-style Markdown tables.
    Returns a list of 2-D lists (rows × cols), separator rows stripped.
    """
    tables = []
    lines  = md_text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith("|") and line.endswith("|"):
            raw = []
            while i < len(lines):
                l = lines[i].strip()
                if not (l.startswith("|") and l.endswith("|")):
                    break
                raw.append(l)
                i += 1
            rows = [
                [c.strip() for c in r.strip("|").split("|")]
                for r in raw
                if not re.match(r"^\s*\|?[\s\-|:]+\|?\s*$", r)
            ]
            if rows:
                tables.append(rows)
        else:
            i += 1
    return tables


def export_tables(md_path: Path, md_text: str, fmt: str, out_dir: Path) -> None:
    """
    Export all tables from *md_text* to CSV or XLSX.

    Parameters
    ----------
    fmt : "csv" | "xlsx"
    """
    tables = extract_tables(md_text)
    if not tables:
        warn("No Markdown tables found.")
        return

    out_dir.mkdir(parents=True, exist_ok=True)
    stem = md_path.stem

    if fmt == "csv":
        for idx, tbl in enumerate(tables, 1):
            dest = out_dir / (stem + "_table" + str(idx) + ".csv")
            with dest.open("w", newline="", encoding="utf-8") as f:
                csv.writer(f).writerows(tbl)
            ok("Saved " + dest.name +
               "  (" + str(len(tbl)) + " rows × " + str(len(tbl[0])) + " cols)")
        return

    # ── XLSX ──────────────────────────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        err("openpyxl not installed.  Run: pip install openpyxl")
        return

    dest = out_dir / (stem + "_tables.xlsx")
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)

    hf   = PatternFill("solid", fgColor="1F3864")
    hfnt = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    af   = PatternFill("solid", fgColor="E8F0FE")
    thin = Side(border_style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, tbl in enumerate(tables, 1):
        ws = wb.create_sheet(title="Table_" + str(idx))
        for ri, row in enumerate(tbl, 1):
            for ci, val in enumerate(row, 1):
                cell           = ws.cell(row=ri, column=ci, value=val)
                cell.border    = bdr
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                if ri == 1:
                    cell.font = hfnt
                    cell.fill = hf
                elif ri % 2 == 0:
                    cell.fill = af
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 50)
        ws.row_dimensions[1].height = 20

    wb.save(dest)
    ok("Saved " + dest.name + "  (" + str(len(tables)) + " sheet(s))")
