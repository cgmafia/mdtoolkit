#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════╗
║            Markdown Toolkit  —  mdtoolkit                        ║
╠══════════════════════════════════════════════════════════════════╣
║  Author  :  Anand Venkataraman <vand3dup@gmail.com>              ║
║  License :  MIT                                                   ║
║  Version :  1.0.0                                                 ║
║  GitHub  :  https://github.com/anandvenkataraman/mdtoolkit        ║
╠══════════════════════════════════════════════════════════════════╣
║  A CLI toolkit to extract code, export tables, generate styled   ║
║  HTML/PDF from Markdown files — with GitHub repository support.  ║
╚══════════════════════════════════════════════════════════════════╝

Usage:
  python3 tool.py <filename.md>      # process a local file
  python3 tool.py                    # interactive launcher (GitHub mode)

Copyright (c) 2026 Anand Venkataraman
Released under the MIT License — see LICENSE file for details.
"""

__author__  = "Anand Venkataraman"
__email__   = "vand3dup@gmail.com"
__version__ = "1.0.0"
__license__ = "MIT"

import sys
import re
import csv
import subprocess
import shutil
import tempfile
import zipfile
import urllib.request
import urllib.error
import os
from pathlib import Path

# ── ANSI helpers ─────────────────────────────────────────────────────────────
R   = "\033[0;31m"
G   = "\033[0;32m"
Y   = "\033[0;33m"
B   = "\033[0;34m"
C   = "\033[0;36m"
W   = "\033[1;37m"
M   = "\033[0;35m"
DIM = "\033[2m"
RST = "\033[0m"


def banner():
    print(B + "╔══════════════════════════════════════════════════════════╗")
    print("║  " + W + "📄  M A R K D O W N   T O O L K I T  v1.0.0" + B + "              ║")
    print("║  " + DIM + "Parse • Extract Code • Export Tables • PDF • GitHub" + B + "     ║")
    print("║  " + DIM + "Author: Anand Venkataraman <vand3dup@gmail.com>" + B + "         ║")
    print("╚══════════════════════════════════════════════════════════╝" + RST)
    print()


def info(msg):  print("  " + C + "ℹ  " + RST + str(msg))
def ok(msg):    print("  " + G + "✔  " + RST + str(msg))
def warn(msg):  print("  " + Y + "⚠  " + RST + str(msg))
def err(msg):   print("  " + R + "✖  " + RST + str(msg))
def head(msg):  print("\n" + W + str(msg) + RST + "\n" + "─" * 52)
def dim(msg):   print("  " + DIM + str(msg) + RST)
def step(msg):  print("  " + M + "▶  " + RST + str(msg))


# ── Load file ─────────────────────────────────────────────────────────────────
def load_md(path: Path) -> str:
    if not path.exists():
        err("File not found: " + str(path))
        sys.exit(1)
    return path.read_text(encoding="utf-8")


# ─────────────────────────────────────────────────────────────────────────────
#  GITHUB FEATURE
# ─────────────────────────────────────────────────────────────────────────────

def parse_github_url(raw: str):
    """
    Accept any of these formats and return (owner, repo, branch_hint):
      https://github.com/owner/repo
      https://github.com/owner/repo.git
      https://github.com/owner/repo/tree/branch
      git@github.com:owner/repo.git
      owner/repo   (shorthand)

    branch_hint is None when not present in the URL.
    """
    raw = raw.strip().rstrip("/")

    # git SSH  →  git@github.com:owner/repo.git
    ssh = re.match(r"git@github\.com:([^/]+)/([^/]+?)(?:\.git)?$", raw)
    if ssh:
        return ssh.group(1), ssh.group(2), None

    # Full HTTPS with optional /tree/<branch>
    https = re.match(
        r"https?://github\.com/([^/]+)/([^/]+?)(?:\.git)?(?:/tree/(.+))?$",
        raw,
    )
    if https:
        return https.group(1), https.group(2), https.group(3)

    # Shorthand  owner/repo
    short = re.match(r"^([A-Za-z0-9_.-]+)/([A-Za-z0-9_.-]+)$", raw)
    if short:
        return short.group(1), short.group(2), None

    return None, None, None


def clone_with_git(clone_url: str, dest: Path, branch: str = None) -> bool:
    """Try git clone. Returns True on success."""
    cmd = ["git", "clone", "--depth", "1", "--quiet"]
    if branch:
        cmd += ["--branch", branch]
    cmd += [clone_url, str(dest)]
    step("git clone " + clone_url + (" @ " + branch if branch else ""))
    r = subprocess.run(cmd, capture_output=True)
    if r.returncode == 0:
        ok("Clone complete")
        return True
    err("git clone failed: " + r.stderr.decode()[:300].strip())
    return False


def download_zip(owner: str, repo: str, branch: str, dest: Path) -> bool:
    """
    Download GitHub ZIP archive via HTTPS API.
    Falls back through main → master → HEAD if branch not specified.
    Returns True on success.
    """
    branches_to_try = [branch] if branch else ["main", "master", "HEAD"]
    for br in branches_to_try:
        url = "https://github.com/" + owner + "/" + repo + "/archive/refs/heads/" + br + ".zip"
        if br == "HEAD":
            url = "https://github.com/" + owner + "/" + repo + "/archive/HEAD.zip"
        step("Trying ZIP download: " + url)
        try:
            tmp_zip = dest.parent / (repo + "_" + br + ".zip")
            urllib.request.urlretrieve(url, str(tmp_zip))
            # Unzip
            with zipfile.ZipFile(str(tmp_zip), "r") as zf:
                zf.extractall(str(dest.parent))
            tmp_zip.unlink()
            # GitHub names the extracted folder  repo-branch/
            extracted = dest.parent / (repo + "-" + br)
            if extracted.exists():
                if dest.exists():
                    shutil.rmtree(dest)
                extracted.rename(dest)
            ok("ZIP download complete (" + br + " branch)")
            return True
        except urllib.error.HTTPError as e:
            warn("HTTP " + str(e.code) + " for branch '" + br + "' — trying next")
        except Exception as e:
            warn("ZIP error: " + str(e))
    return False


def find_md_files(root: Path) -> list:
    """
    Recursively find all .md / .markdown files under *root*.
    Skips hidden directories (.git, .github, node_modules, vendor …).
    Returns sorted list of Path objects, README.md first.
    """
    skip_dirs = {".git", ".github", "node_modules", "vendor", ".tox",
                 "__pycache__", "venv", ".venv", "dist", "build"}
    found = []
    for p in root.rglob("*"):
        if any(part in skip_dirs for part in p.parts):
            continue
        if p.suffix.lower() in (".md", ".markdown") and p.is_file():
            found.append(p)

    # Sort: README.md first, then alphabetical by relative path
    def sort_key(p):
        rel = p.relative_to(root)
        is_readme = p.stem.lower() == "readme"
        return (0 if is_readme else 1, str(rel).lower())

    return sorted(found, key=sort_key)


def pick_md_file(md_files: list, repo_root: Path) -> Path:
    """
    Interactive picker.  If only one file exists, return it directly.
    Otherwise show a numbered list and let the user choose.
    Returns the chosen Path, or None to cancel.
    """
    if not md_files:
        err("No Markdown files found in this repository.")
        return None

    if len(md_files) == 1:
        info("Single Markdown file found: " + str(md_files[0].relative_to(repo_root)))
        return md_files[0]

    head("📑  Markdown files found in repository")
    for i, p in enumerate(md_files, 1):
        rel  = p.relative_to(repo_root)
        size = p.stat().st_size
        size_str = (str(size // 1024) + " KB") if size >= 1024 else (str(size) + "  B")
        print("  " + B + "[" + str(i).rjust(2) + "]" + RST +
              "  " + W + str(rel) + RST +
              "  " + DIM + size_str + RST)

    print("  " + B + "[ A]" + RST + "  " + Y + "Process ALL files" + RST)
    print("  " + B + "[ 0]" + RST + "  " + DIM + "Cancel / go back" + RST)

    while True:
        raw = input("\n  " + W + "Choose file (number / A / 0): " + RST).strip().upper()
        if raw == "0":
            return None
        if raw == "A":
            return "ALL"
        try:
            idx = int(raw)
            if 1 <= idx <= len(md_files):
                return md_files[idx - 1]
            warn("Enter a number between 1 and " + str(len(md_files)))
        except ValueError:
            warn("Invalid input — enter a number, A, or 0")


def github_flow(base_out_dir: Path):
    """
    Full GitHub → download → pick MD → run menu flow.
    base_out_dir: where clones and outputs will be written.
    """
    head("🐙  GitHub Repository Mode")

    # ── Get URL ──────────────────────────────────────────────────────────────
    print()
    print("  " + DIM + "Examples:" + RST)
    print("  " + DIM + "  https://github.com/owner/repo" + RST)
    print("  " + DIM + "  https://github.com/owner/repo/tree/develop" + RST)
    print("  " + DIM + "  git@github.com:owner/repo.git" + RST)
    print("  " + DIM + "  owner/repo" + RST)
    print()

    raw_url = input("  " + W + "GitHub URL or shorthand: " + RST).strip()
    if not raw_url:
        warn("No URL entered — returning to menu.")
        return

    owner, repo, branch_hint = parse_github_url(raw_url)
    if not owner:
        err("Could not parse GitHub URL: " + raw_url)
        info("Expected format: https://github.com/owner/repo  or  owner/repo")
        return

    info("Owner  : " + owner)
    info("Repo   : " + repo)
    info("Branch : " + (branch_hint or "(auto-detect)"))

    # ── Optional branch override ──────────────────────────────────────────────
    if not branch_hint:
        override = input(
            "\n  " + DIM + "Branch name (leave blank for default main/master): " + RST
        ).strip()
        if override:
            branch_hint = override

    # ── Destination directory ─────────────────────────────────────────────────
    clone_dir = base_out_dir / "repos" / (owner + "_" + repo)
    if clone_dir.exists():
        print()
        warn("Local copy already exists: " + str(clone_dir))
        reuse = input("  " + W + "Use cached copy? [Y/n]: " + RST).strip().lower()
        if reuse not in ("n", "no"):
            info("Reusing cached repository.")
        else:
            step("Removing old copy …")
            shutil.rmtree(clone_dir)
            clone_dir.mkdir(parents=True, exist_ok=True)
            _do_download(owner, repo, branch_hint, clone_dir)
    else:
        clone_dir.mkdir(parents=True, exist_ok=True)
        success = _do_download(owner, repo, branch_hint, clone_dir)
        if not success:
            shutil.rmtree(clone_dir, ignore_errors=True)
            return

    # ── Discover MD files ─────────────────────────────────────────────────────
    step("Scanning repository for Markdown files …")
    md_files = find_md_files(clone_dir)
    info("Found " + str(len(md_files)) + " Markdown file(s)")

    while True:
        choice = pick_md_file(md_files, clone_dir)

        if choice is None:
            info("Cancelled — returning to main menu.")
            return

        if choice == "ALL":
            # Process every MD file in sequence
            for md_path in md_files:
                out_dir = base_out_dir / "github_output" / (owner + "_" + repo) / md_path.stem
                print()
                info("Processing: " + str(md_path.relative_to(clone_dir)))
                run_menu(md_path, out_dir, single_shot=False)
            info("All files processed.")
            return
        else:
            # Single file — run interactive menu then offer to pick another
            out_dir = base_out_dir / "github_output" / (owner + "_" + repo) / choice.stem
            run_menu(choice, out_dir)
            print()
            again = input("  " + W + "Process another file from this repo? [y/N]: " + RST).strip().lower()
            if again not in ("y", "yes"):
                return


def _do_download(owner: str, repo: str, branch: str, dest: Path) -> bool:
    """Try git clone first, fall back to ZIP download."""
    clone_url = "https://github.com/" + owner + "/" + repo + ".git"

    # 1. Try git clone
    if clone_with_git(clone_url, dest, branch):
        return True

    # 2. Fall back to ZIP download
    warn("git clone failed — trying ZIP download …")
    if download_zip(owner, repo, branch, dest):
        return True

    err("Could not download repository.")
    info("Check the URL, your network connection, and that the repo is public.")
    return False


# ── 1.  CODE EXTRACTION ───────────────────────────────────────────────────────

EXT_MAP = {
    "python": "py",    "py": "py",        "java": "java",
    "javascript": "js","js": "js",        "typescript": "ts",  "ts": "ts",
    "sql": "sql",      "bash": "sh",      "shell": "sh",       "sh": "sh",
    "html": "html",    "css": "css",      "xml": "xml",
    "yaml": "yml",     "yml": "yml",      "json": "json",
    "kotlin": "kt",    "go": "go",        "cpp": "cpp",        "c": "c",
    "ruby": "rb",      "rust": "rs",      "php": "php",
    "swift": "swift",  "scala": "scala",  "r": "r",
    "markdown": "md",  "md": "md",        "dockerfile": "dockerfile",
}


def extract_code_blocks(md_text: str):
    """
    Returns list of (lang, hint, code, line_no).
    line_no = 1-based line of the opening ``` fence → used in fallback filename.
    """
    fence_re = re.compile(
        r"```[ \t]*([a-zA-Z0-9_+#.-]*)(?::([^\n]+))?\n(.*?)```",
        re.DOTALL,
    )
    line_starts = [0]
    for i, ch in enumerate(md_text):
        if ch == "\n":
            line_starts.append(i + 1)

    def offset_to_line(offset):
        lo, hi = 0, len(line_starts) - 1
        while lo < hi:
            mid = (lo + hi + 1) // 2
            if line_starts[mid] <= offset:
                lo = mid
            else:
                hi = mid - 1
        return lo + 1

    result = []
    for m in fence_re.finditer(md_text):
        lang    = m.group(1).strip() if m.group(1) else ""
        hint    = m.group(2).strip() if m.group(2) else ""
        code    = m.group(3)
        line_no = offset_to_line(m.start())
        result.append((lang, hint, code, line_no))
    return result


def save_code_files(md_path: Path, md_text: str, out_dir: Path):
    blocks = extract_code_blocks(md_text)
    if not blocks:
        warn("No fenced code blocks found.")
        return

    out_dir.mkdir(parents=True, exist_ok=True)
    saved = []

    for lang, hint, code, line_no in blocks:
        ext = EXT_MAP.get(lang.lower(), "txt") if lang else "txt"
        fname = hint.replace("/", "_") if hint else ("code_line" + str(line_no) + "." + ext)
        dest  = out_dir / fname
        stem, sfx = dest.stem, dest.suffix
        dup = 1
        while dest.exists():
            dest = out_dir / (stem + "_" + str(dup) + sfx)
            dup += 1
        dest.write_text(code, encoding="utf-8")
        saved.append((dest.name, lang or "—", line_no, len(code.splitlines())))

    h1, h2, h3, h4 = "FILE", "LANG", "LINE", "LINES"
    print()
    print("  " + G + h1.ljust(35) + h2.ljust(16) + h3.rjust(5) + "  " + h4 + RST)
    print("  " + "─" * 64)
    for fname, lang, lno, cnt in saved:
        print("  " + W + fname.ljust(35) + RST +
              C + lang.ljust(16) + RST +
              DIM + str(lno).rjust(5) + RST + "  " + str(cnt))
    print()
    ok(str(len(saved)) + " file(s) saved to  " + str(out_dir) + "/")


# ── 2.  TABLE EXTRACTION ──────────────────────────────────────────────────────

def extract_tables(md_text: str):
    tables = []
    lines  = md_text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith("|") and line.endswith("|"):
            raw = []
            while i < len(lines):
                l = lines[i].strip()
                if not (l.startswith("|") and l.endswith("|")):
                    break
                raw.append(l)
                i += 1
            rows = [
                [c.strip() for c in r.strip("|").split("|")]
                for r in raw
                if not re.match(r"^\s*\|?[\s\-|:]+\|?\s*$", r)
            ]
            if rows:
                tables.append(rows)
        else:
            i += 1
    return tables


def export_tables(md_path: Path, md_text: str, fmt: str, out_dir: Path):
    tables = extract_tables(md_text)
    if not tables:
        warn("No Markdown tables found.")
        return

    out_dir.mkdir(parents=True, exist_ok=True)
    stem = md_path.stem

    if fmt == "csv":
        for idx, tbl in enumerate(tables, 1):
            dest = out_dir / (stem + "_table" + str(idx) + ".csv")
            with dest.open("w", newline="", encoding="utf-8") as f:
                csv.writer(f).writerows(tbl)
            ok("Saved " + dest.name + "  (" + str(len(tbl)) +
               " rows × " + str(len(tbl[0])) + " cols)")
        return

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        err("openpyxl not installed.  Run: pip install openpyxl")
        return

    dest = out_dir / (stem + "_tables.xlsx")
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)
    hf   = PatternFill("solid", fgColor="1F3864")
    hfnt = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    af   = PatternFill("solid", fgColor="E8F0FE")
    thin = Side(border_style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, tbl in enumerate(tables, 1):
        ws = wb.create_sheet(title="Table_" + str(idx))
        for ri, row in enumerate(tbl, 1):
            for ci, val in enumerate(row, 1):
                cell           = ws.cell(row=ri, column=ci, value=val)
                cell.border    = bdr
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                if ri == 1:
                    cell.font = hfnt;  cell.fill = hf
                elif ri % 2 == 0:
                    cell.fill = af
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 50)
        ws.row_dimensions[1].height = 20

    wb.save(dest)
    ok("Saved " + dest.name + "  (" + str(len(tables)) + " sheet(s))")


# ── 3.  STYLED HTML ───────────────────────────────────────────────────────────

CSS = """
:root{--bg:#0f172a;--surface:#1e293b;--border:#334155;--accent:#3b82f6;
  --accent2:#8b5cf6;--text:#e2e8f0;--muted:#94a3b8;--code-bg:#020617;
  --link:#60a5fa;--font:'Segoe UI',system-ui,sans-serif;
  --mono:'JetBrains Mono','Fira Code',monospace}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--text);font-family:var(--font);
  font-size:16px;line-height:1.75}
.md-container{max-width:900px;margin:0 auto;padding:60px 40px}
h1,h2,h3,h4,h5,h6{font-weight:700;line-height:1.3;margin:2em 0 .6em;letter-spacing:-.02em}
h1{font-size:2.4em;background:linear-gradient(135deg,#60a5fa,#a78bfa);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;
  background-clip:text;border-bottom:2px solid var(--border);padding-bottom:.4em}
h2{font-size:1.7em;color:#93c5fd;border-bottom:1px solid var(--border);padding-bottom:.3em}
h3{font-size:1.3em;color:#a5b4fc}
h4{font-size:1.1em;color:#c4b5fd}
p{margin:.9em 0}
a{color:var(--link);text-decoration:none;border-bottom:1px solid #3b82f640}
a:hover{color:#93c5fd;border-color:#93c5fd}
code{font-family:var(--mono);font-size:.85em;background:var(--code-bg);color:#f472b6;
  padding:2px 6px;border-radius:4px;border:1px solid var(--border)}
pre{background:var(--code-bg);border:1px solid var(--border);
  border-left:4px solid var(--accent);border-radius:8px;
  padding:1.2em 1.5em;overflow-x:auto;margin:1.5em 0;position:relative}
pre code{background:transparent;color:var(--text);border:none;padding:0;
  font-size:.88em;line-height:1.6}
pre::before{content:attr(data-lang);position:absolute;top:8px;right:14px;
  font-size:.72em;color:var(--muted);text-transform:uppercase;
  letter-spacing:.08em;font-family:var(--font)}
table{width:100%;border-collapse:collapse;margin:1.8em 0;
  border-radius:8px;overflow:hidden;box-shadow:0 0 0 1px var(--border)}
thead tr{background:linear-gradient(135deg,#1e3a5f,#2d1f6e)}
th{color:#e2e8f0;font-weight:600;font-size:.85em;text-transform:uppercase;
  letter-spacing:.06em;padding:12px 16px;text-align:left}
td{padding:11px 16px;border-top:1px solid var(--border);font-size:.94em}
tbody tr:nth-child(even){background:#162032}
tbody tr:hover{background:#1a2840}
blockquote{border-left:4px solid var(--accent2);background:#1e1b4b40;
  padding:.8em 1.2em;border-radius:0 8px 8px 0;margin:1.5em 0;
  color:var(--muted);font-style:italic}
ul,ol{padding-left:1.6em;margin:.8em 0}
li{margin:.35em 0}
ul li::marker{color:var(--accent)}
ol li::marker{color:var(--accent2);font-weight:700}
hr{border:none;border-top:1px solid var(--border);margin:2.5em 0}
strong{color:#f8fafc;font-weight:700}
em{color:#c4b5fd}
"""


def build_html(md_path: Path, md_text: str) -> str:
    try:
        import markdown as md_lib
        html_body = md_lib.markdown(
            md_text,
            extensions=["fenced_code", "tables", "toc", "attr_list", "nl2br"],
        )
    except ImportError:
        html_body = _minimal_md_to_html(md_text)

    def tag_pre(m):
        cls  = m.group(1)
        body = m.group(2)
        lang = re.search(r'class="language-(\w+)"', cls)
        dl   = ' data-lang="' + lang.group(1) + '"' if lang else ""
        return "<pre" + dl + "><code " + cls + ">" + body + "</code></pre>"

    html_body = re.sub(r'<pre><code ([^>]*)>(.*?)</code></pre>',
                       tag_pre, html_body, flags=re.DOTALL)

    title = md_path.stem.replace("_", " ").replace("-", " ").title()
    return (
        '<!DOCTYPE html>\n<html lang="en">\n<head>\n'
        '  <meta charset="UTF-8"/>\n'
        '  <meta name="viewport" content="width=device-width,initial-scale=1.0"/>\n'
        '  <title>' + title + '</title>\n'
        '  <style>\n' + CSS + '\n  </style>\n</head>\n<body>\n'
        '  <div class="md-container">\n' + html_body + '\n  </div>\n</body>\n</html>\n'
    )


def _minimal_md_to_html(text: str) -> str:
    t = text
    t = re.sub(r"^# (.+)$",   r"<h1>\1</h1>", t, flags=re.M)
    t = re.sub(r"^## (.+)$",  r"<h2>\1</h2>", t, flags=re.M)
    t = re.sub(r"^### (.+)$", r"<h3>\1</h3>", t, flags=re.M)
    t = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", t)
    t = re.sub(r"\*(.+?)\*",     r"<em>\1</em>",         t)
    t = re.sub(r"`(.+?)`",       r"<code>\1</code>",      t)
    t = re.sub(r"```.*?\n(.*?)```", r"<pre><code>\1</code></pre>", t, flags=re.DOTALL)
    return "\n".join(
        ("<p>" + l + "</p>" if not re.match(r"^<", l.strip()) and l.strip() else l)
        for l in t.splitlines()
    )


def save_html(md_path: Path, md_text: str, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    dest = out_dir / (md_path.stem + ".html")
    dest.write_text(build_html(md_path, md_text), encoding="utf-8")
    return dest


# ── 4.  PDF EXPORT ────────────────────────────────────────────────────────────

def export_pdf(md_path: Path, md_text: str, out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)
    html_path = save_html(md_path, md_text, out_dir)
    pdf_path  = out_dir / (md_path.stem + ".pdf")

    if subprocess.run(["which", "wkhtmltopdf"], capture_output=True).returncode == 0:
        r = subprocess.run(
            ["wkhtmltopdf", "--page-size", "A4",
             "--margin-top", "15mm", "--margin-bottom", "15mm",
             "--margin-left", "12mm", "--margin-right", "12mm",
             "--enable-local-file-access", "--quiet",
             str(html_path), str(pdf_path)],
            capture_output=True,
        )
        if r.returncode == 0:
            ok("PDF saved  →  " + str(pdf_path));  return
        warn("wkhtmltopdf: " + r.stderr.decode()[:200])

    for browser in ("google-chrome", "chromium-browser", "chromium"):
        if subprocess.run(["which", browser], capture_output=True).returncode == 0:
            r = subprocess.run(
                [browser, "--headless", "--disable-gpu", "--no-sandbox",
                 "--print-to-pdf=" + str(pdf_path), str(html_path)],
                capture_output=True,
            )
            if r.returncode == 0:
                ok("PDF saved  →  " + str(pdf_path));  return
            warn(browser + ": " + r.stderr.decode()[:200])

    err("No PDF renderer found (wkhtmltopdf or Chrome/Chromium required).")
    info("HTML exported instead → " + str(html_path))


# ── MENUS ─────────────────────────────────────────────────────────────────────

MENU_ITEMS = [
    ("1", "🗂  Extract Code Blocks",   "named from markdown or code_line<N>.ext"),
    ("2", "📊  Export Tables -> CSV",  "one .csv per table"),
    ("3", "📊  Export Tables -> XLSX", "all tables in one styled workbook"),
    ("4", "🎨  Export Styled HTML",    "dark-themed HTML with embedded CSS"),
    ("5", "📄  Export PDF",            "via wkhtmltopdf or Chrome headless"),
    ("6", "🚀  Do Everything",         "run all 5 actions at once"),
    ("7", "🔍  File Summary",          "headings, code blocks, tables count"),
    ("Q", "🚪  Quit / Back",           ""),
]

LAUNCH_MENU = [
    ("1", "📂  Open local Markdown file",     "pass a path or enter it now"),
    ("2", "🐙  GitHub repository",            "clone & pick an MD file"),
    ("Q", "🚪  Quit",                         ""),
]


def print_menu():
    print("\n" + W + "  ┌─ What would you like to do? ─────────────────────────┐" + RST)
    for key, label, hint in MENU_ITEMS:
        hint_str = ("  " + DIM + hint + RST) if hint else ""
        print("  " + B + "[" + key + "]" + RST + "  " + label + hint_str)
    print(W + "  └──────────────────────────────────────────────────────────┘" + RST)


def print_launch_menu():
    print("\n" + W + "  ┌─ Choose mode ───────────────────────────────────────────┐" + RST)
    for key, label, hint in LAUNCH_MENU:
        hint_str = ("  " + DIM + hint + RST) if hint else ""
        print("  " + B + "[" + key + "]" + RST + "  " + label + hint_str)
    print(W + "  └──────────────────────────────────────────────────────────┘" + RST)


def file_summary(md_path: Path, md_text: str):
    head("📋  Summary of  " + md_path.name)
    lines  = md_text.splitlines()
    blocks = extract_code_blocks(md_text)
    tables = extract_tables(md_text)
    links  = re.findall(r"\[([^\]]+)\]\(([^)]+)\)", md_text)

    info("Total lines  : " + str(len(lines)))
    info("Words        : " + str(len(md_text.split())))
    info("H1 headings  : " + str(sum(1 for l in lines if l.startswith("# "))))
    info("H2 headings  : " + str(sum(1 for l in lines if l.startswith("## "))))
    info("H3 headings  : " + str(sum(1 for l in lines if l.startswith("### "))))
    info("Code blocks  : " + str(len(blocks)))
    info("Tables       : " + str(len(tables)))
    info("Links        : " + str(len(links)))

    if blocks:
        head("Code blocks")
        for i, (lang, hint, code, line_no) in enumerate(blocks, 1):
            ext   = EXT_MAP.get(lang.lower(), "txt") if lang else "txt"
            label = hint if hint else "(→ code_line" + str(line_no) + "." + ext + ")"
            dim("  #" + str(i).ljust(3) +
                " line=" + str(line_no).ljust(5) +
                " lang=" + (lang or "none").ljust(12) + "  " + label)

    if tables:
        head("Tables")
        for i, tbl in enumerate(tables, 1):
            dim("  #" + str(i) + "  " + str(len(tbl[0])) + " cols x " +
                str(len(tbl)) + " rows  |  header: " + ", ".join(tbl[0]))


def run_menu(md_path: Path, out_dir: Path, single_shot: bool = False):
    """
    Interactive per-file menu.
    single_shot=True → run once and return (used by batch GitHub processing).
    """
    md_text = load_md(md_path)

    if not single_shot:
        info("File  :  " + str(md_path))
        info("Output:  " + str(out_dir) + "/")

    while True:
        if not single_shot:
            print_menu()
            choice = input("\n  " + W + "Enter choice: " + RST).strip().upper()
        else:
            # In batch mode run everything automatically
            choice = "6"

        if choice == "1":
            head("📂  Extracting Code")
            save_code_files(md_path, md_text, out_dir / "code")
        elif choice == "2":
            head("📊  Exporting Tables -> CSV")
            export_tables(md_path, md_text, "csv", out_dir / "tables")
        elif choice == "3":
            head("📊  Exporting Tables -> XLSX")
            export_tables(md_path, md_text, "xlsx", out_dir / "tables")
        elif choice == "4":
            head("🎨  Exporting Styled HTML")
            ok("HTML saved  ->  " + str(save_html(md_path, md_text, out_dir)))
        elif choice == "5":
            head("📄  Exporting PDF")
            export_pdf(md_path, md_text, out_dir)
        elif choice == "6":
            head("🚀  Running All Actions")
            save_code_files(md_path, md_text, out_dir / "code")
            export_tables(md_path, md_text, "csv",  out_dir / "tables")
            export_tables(md_path, md_text, "xlsx", out_dir / "tables")
            ok("HTML  ->  " + str(save_html(md_path, md_text, out_dir)))
            export_pdf(md_path, md_text, out_dir)
            ok("All done!")
            if single_shot:
                return
        elif choice == "7":
            file_summary(md_path, md_text)
        elif choice in ("Q", "QUIT", "EXIT", "0", "B", "BACK"):
            return
        else:
            warn("Invalid choice. Please try again.")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    banner()

    # ── Called with a file path ───────────────────────────────────────────────
    if len(sys.argv) >= 2:
        md_path = Path(sys.argv[1]).resolve()
        out_dir = md_path.parent / (md_path.stem + "_output")
        run_menu(md_path, out_dir)
        sys.exit(0)

    # ── No arguments → launch menu ────────────────────────────────────────────
    # Default base output directory: current working directory / md_toolkit_output
    base_out = Path.cwd() / "md_toolkit_output"
    base_out.mkdir(parents=True, exist_ok=True)

    while True:
        print_launch_menu()
        choice = input("\n  " + W + "Enter choice: " + RST).strip().upper()

        if choice == "1":
            raw = input("  " + W + "Path to Markdown file: " + RST).strip()
            if not raw:
                warn("No path entered.")
                continue
            md_path = Path(raw).expanduser().resolve()
            if not md_path.exists():
                err("File not found: " + str(md_path))
                continue
            out_dir = base_out / md_path.stem
            run_menu(md_path, out_dir)

        elif choice == "2":
            github_flow(base_out)

        elif choice in ("Q", "QUIT", "EXIT", "0"):
            print("\n  " + G + "Goodbye! ✨" + RST + "\n")
            break
        else:
            warn("Invalid choice.")
